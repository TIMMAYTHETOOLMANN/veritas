#!/usr/bin/env python3
"""
Utility Bill Email Scraper for Gmail
Searches timothyroessel@gmail.com for utility bills and compiles averages.
"""

import os
import re
import json
import pickle
import base64
from datetime import datetime, timedelta
from collections import defaultdict
from dataclasses import dataclass, asdict
from typing import List, Optional

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# Gmail API scopes
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

# Utility categories and search keywords
UTILITY_CATEGORIES = {
    'electric': [
        'electric', 'electricity', 'power bill', 'energy bill',
        'pge', 'pacific gas', 'edison', 'sce', 'sdge', 'con ed',
        'duke energy', 'dominion', 'national grid', 'eversource',
        'kwh', 'kilowatt', 'electric service'
    ],
    'gas': [
        'gas bill', 'natural gas', 'gas service', 'therms',
        'pge gas', 'socal gas', 'nicor', 'peoples gas',
        'gas utility', 'gas company'
    ],
    'water': [
        'water bill', 'water service', 'water utility',
        'water company', 'water dept', 'municipal water',
        'sewer', 'sewage', 'wastewater', 'stormwater',
        'water/sewer', 'water & sewer'
    ],
    'internet': [
        'internet bill', 'internet service', 'broadband',
        'wifi bill', 'cable internet', 'fiber internet',
        'comcast', 'xfinity', 'spectrum', 'verizon fios',
        'att internet', 'at&t internet', 'cox', 'optimum',
        'frontier', 'centurylink', 'windstream', 'google fiber'
    ],
    'phone': [
        'phone bill', 'cell phone', 'mobile bill', 'wireless',
        'verizon wireless', 'att wireless', 't-mobile',
        'sprint', 'mint mobile', 'visible', 'cricket',
        'landline', 'home phone'
    ],
    'trash': [
        'trash bill', 'garbage bill', 'waste management',
        'trash service', 'garbage service', 'recycling',
        'republic services', 'waste connections', 'wm.com'
    ],
    'hoa': [
        'hoa', 'homeowners association', 'condo fee',
        'association dues', 'community association'
    ],
    'other': [
        'utility bill', 'utilities', 'utility service'
    ]
}

@dataclass
class BillRecord:
    category: str
    date: str
    amount: float
    sender: str
    subject: str
    snippet: str

def get_gmail_service():
    """Authenticate and return Gmail API service."""
    creds = None
    token_path = 'token.pickle'
    
    if os.path.exists(token_path):
        with open(token_path, 'rb') as token:
            creds = pickle.load(token)
    
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            # You'll need credentials.json from Google Cloud Console
            if not os.path.exists('credentials.json'):
                print("ERROR: credentials.json not found!")
                print("Download from Google Cloud Console > APIs & Services > Credentials")
                print("Save as 'credentials.json' in this directory")
                return None
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        
        with open(token_path, 'wb') as token:
            pickle.dump(creds, token)
    
    return build('gmail', 'v1', credentials=creds)

def categorize_email(subject: str, snippet: str, sender: str) -> Optional[str]:
    """Determine utility category from email content."""
    text = f"{subject} {snippet} {sender}".lower()
    
    for category, keywords in UTILITY_CATEGORIES.items():
        for keyword in keywords:
            if keyword.lower() in text:
                return category
    return None

def extract_amount(text: str) -> Optional[float]:
    """Extract dollar amount from text."""
    # Patterns for dollar amounts
    patterns = [
        r'\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',  # $123.45 or $1,234.56
        r'(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*dollars?',
        r'amount\s*(?:due|:|\$)\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'balance\s*(?:due|:|\$)\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'total\s*(?:due|:|\$)\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
        r'current\s*(?:charges?|amount)\s*(?:due|:|\$)\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        if matches:
            # Take the largest amount found (likely the total)
            amounts = [float(m.replace(',', '')) for m in matches]
            if amounts:
                return max(amounts)
    return None

def parse_date(date_str: str) -> Optional[str]:
    """Parse Gmail date header to YYYY-MM-DD."""
    try:
        # Gmail dates like: "Wed, 15 Jan 2025 10:30:00 -0500"
        dt = datetime.strptime(date_str[:31], '%a, %d %b %Y %H:%M:%S')
        return dt.strftime('%Y-%m-%d')
    except:
        try:
            dt = datetime.strptime(date_str[:25], '%d %b %Y %H:%M:%S')
            return dt.strftime('%Y-%m-%d')
        except:
            return None

def search_utility_emails(service, max_results: int = 500) -> List[dict]:
    """Search Gmail for utility-related emails."""
    # Build query with utility keywords
    keywords = []
    for cat_keywords in UTILITY_CATEGORIES.values():
        keywords.extend(cat_keywords)
    
    # Create OR query for subject/body
    query_parts = []
    for kw in keywords[:30]:  # Limit query length
        query_parts.append(f'subject:"{kw}"')
        query_parts.append(f'body:"{kw}"')
    
    query = ' OR '.join(query_parts)
    query += ' newer_than:2y'  # Last 2 years
    
    print(f"Searching with query ({len(query)} chars)...")
    
    try:
        results = service.users().messages().list(
            userId='me',
            q=query,
            maxResults=max_results
        ).execute()
        
        messages = results.get('messages', [])
        print(f"Found {len(messages)} messages")
        
        # Fetch full messages
        emails = []
        for i, msg in enumerate(messages):
            if i % 50 == 0:
                print(f"  Fetching {i+1}/{len(messages)}...")
            
            full_msg = service.users().messages().get(
                userId='me',
                id=msg['id'],
                format='full'
            ).execute()
            emails.append(full_msg)
        
        return emails
    
    except HttpError as error:
        print(f"Gmail API error: {error}")
        return []

def process_emails(emails: List[dict]) -> List[BillRecord]:
    """Extract bill records from emails."""
    records = []
    
    for email in emails:
        payload = email.get('payload', {})
        headers = {h['name']: h['value'] for h in payload.get('headers', [])}
        
        subject = headers.get('Subject', '')
        sender = headers.get('From', '')
        date_str = headers.get('Date', '')
        snippet = email.get('snippet', '')
        
        # Get body text
        body_text = ''
        if 'parts' in payload:
            for part in payload['parts']:
                if part['mimeType'] == 'text/plain' and 'data' in part.get('body', {}):
                    body_text = base64.urlsafe_b64decode(part['body']['data']).decode('utf-8', errors='ignore')
                    break
        elif payload.get('body', {}).get('data'):
            body_text = base64.urlsafe_b64decode(payload['body']['data']).decode('utf-8', errors='ignore')
        
        # Categorize
        category = categorize_email(subject, snippet + ' ' + body_text[:2000], sender)
        if not category:
            continue
        
        # Extract amount
        amount = extract_amount(subject + ' ' + snippet + ' ' + body_text[:3000])
        if not amount or amount < 5 or amount > 5000:  # Sanity check
            continue
        
        # Parse date
        date = parse_date(date_str)
        if not date:
            continue
        
        records.append(BillRecord(
            category=category,
            date=date,
            amount=round(amount, 2),
            sender=sender[:100],
            subject=subject[:150],
            snippet=snippet[:200]
        ))
    
    return records

def compute_averages(records: List[BillRecord]) -> dict:
    """Compute monthly and overall averages per category."""
    by_category = defaultdict(list)
    by_category_month = defaultdict(lambda: defaultdict(list))
    
    for r in records:
        by_category[r.category].append(r.amount)
        # Group by year-month
        ym = r.date[:7]  # YYYY-MM
        by_category_month[r.category][ym].append(r.amount)
    
    results = {}
    for cat, amounts in by_category.items():
        monthly_avg = {}
        for ym, month_amounts in by_category_month[cat].items():
            monthly_avg[ym] = round(sum(month_amounts) / len(month_amounts), 2)
        
        results[cat] = {
            'total_bills': len(amounts),
            'overall_avg': round(sum(amounts) / len(amounts), 2),
            'min': round(min(amounts), 2),
            'max': round(max(amounts), 2),
            'monthly_averages': monthly_avg,
            'records': [asdict(r) for r in sorted(
                [rec for rec in records if rec.category == cat],
                key=lambda x: x.date
            )]
        }
    
    return results

def export_to_excel(results: dict, output_path: str):
    """Export results to Excel with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Installing openpyxl...")
        import subprocess
        subprocess.run(['pip', 'install', 'openpyxl'], check=True)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    headers = ['Category', 'Total Bills', 'Overall Avg', 'Min', 'Max', 'Latest Month', 'Latest Avg']
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    row = 2
    for cat, data in sorted(results.items()):
        latest_month = max(data['monthly_averages'].keys()) if data['monthly_averages'] else 'N/A'
        latest_avg = data['monthly_averages'].get(latest_month, 'N/A')
        
        ws_summary.cell(row=row, column=1, value=cat.capitalize()).border = thin_border
        ws_summary.cell(row=row, column=2, value=data['total_bills']).border = thin_border
        ws_summary.cell(row=row, column=3, value=f"${data['overall_avg']:.2f}").border = thin_border
        ws_summary.cell(row=row, column=4, value=f"${data['min']:.2f}").border = thin_border
        ws_summary.cell(row=row, column=5, value=f"${data['max']:.2f}").border = thin_border
        ws_summary.cell(row=row, column=6, value=latest_month).border = thin_border
        ws_summary.cell(row=row, column=7, value=f"${latest_avg:.2f}" if isinstance(latest_avg, float) else latest_avg).border = thin_border
        row += 1
    
    # Adjust column widths
    for col in range(1, 8):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 2: Monthly Trends
    ws_monthly = wb.create_sheet("Monthly Trends")
    
    # Collect all months
    all_months = set()
    for data in results.values():
        all_months.update(data['monthly_averages'].keys())
    all_months = sorted(all_months)
    
    headers = ['Category'] + all_months
    for col, h in enumerate(headers, 1):
        cell = ws_monthly.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, (cat, data) in enumerate(sorted(results.items()), 2):
        ws_monthly.cell(row=row_idx, column=1, value=cat.capitalize()).border = thin_border
        for col_idx, ym in enumerate(all_months, 2):
            val = data['monthly_averages'].get(ym)
            ws_monthly.cell(row=row_idx, column=col_idx, value=f"${val:.2f}" if val else '').border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws_monthly.column_dimensions[get_column_letter(col)].width = 14
    
    # Sheet 3+: Detail per category
    for cat, data in sorted(results.items()):
        ws = wb.create_sheet(cat.capitalize()[:31])
        
        headers = ['Date', 'Amount', 'Sender', 'Subject', 'Snippet']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx, rec in enumerate(data['records'], 2):
            ws.cell(row=row_idx, column=1, value=rec['date']).border = thin_border
            ws.cell(row=row_idx, column=2, value=f"${rec['amount']:.2f}").border = thin_border
            ws.cell(row=row_idx, column=3, value=rec['sender']).border = thin_border
            ws.cell(row=row_idx, column=4, value=rec['subject']).border = thin_border
            ws.cell(row=row_idx, column=5, value=rec['snippet']).border = thin_border
        
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 50
    
    wb.save(output_path)
    print(f"Excel saved to {output_path}")

def main():
    print("=" * 60)
    print("UTILITY BILL EMAIL SCRAPER")
    print("=" * 60)
    
    # Change to script directory
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    
    # Check for credentials
    if not os.path.exists('credentials.json'):
        print("\n⚠️  SETUP REQUIRED:")
        print("1. Go to https://console.cloud.google.com/")
        print("2. Create a project or select existing")
        print("3. Enable Gmail API")
        print("4. Create OAuth 2.0 Client ID (Desktop app)")
        print("5. Download credentials.json to this folder:")
        print(f"   {os.getcwd()}/credentials.json")
        print("\nThen re-run this script.")
        return
    
    # Authenticate
    print("\n🔐 Authenticating with Gmail...")
    service = get_gmail_service()
    if not service:
        return
    
    # Search emails
    print("\n📧 Searching for utility bills...")
    emails = search_utility_emails(service, max_results=500)
    
    if not emails:
        print("No utility emails found. Try adjusting search terms.")
        return
    
    # Process
    print("\n🔍 Processing emails...")
    records = process_emails(emails)
    
    if not records:
        print("No bill amounts extracted. Check categories/keywords.")
        return
    
    print(f"\n✅ Found {len(records)} bill records")
    
    # Compute averages
    print("\n📊 Computing averages...")
    results = compute_averages(records)
    
    # Print summary
    print("\n" + "=" * 60)
    print("RESULTS SUMMARY")
    print("=" * 60)
    for cat, data in sorted(results.items()):
        print(f"\n{cat.upper()}:")
        print(f"  Bills found: {data['total_bills']}")
        print(f"  Average: ${data['overall_avg']:.2f}")
        print(f"  Range: ${data['min']:.2f} - ${data['max']:.2f}")
        print(f"  Monthly averages:")
        for ym, avg in sorted(data['monthly_averages'].items())[-6:]:
            print(f"    {ym}: ${avg:.2f}")
    
    # Export
    output_path = os.path.join(os.getcwd(), 'utility_bills_analysis.xlsx')
    export_to_excel(results, output_path)
    
    # Also save raw JSON
    json_path = os.path.join(os.getcwd(), 'utility_bills_raw.json')
    with open(json_path, 'w') as f:
        json.dump({cat: data for cat, data in results.items()}, f, indent=2)
    print(f"\n📁 Raw data saved to {json_path}")
    
    print("\n✅ Done!")

if __name__ == '__main__':
    main()